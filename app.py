from flask import Flask, request, render_template, jsonify
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import uuid
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Allowed extensions
ALLOWED_IMAGES = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
ALLOWED_VIDEOS = {'mp4', 'mov', 'avi', 'mkv', 'webm'}

def allowed_file(filename, allowed_set):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_set

@app.route('/')
def form():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    try:
        # Collect all form data
        pm_start_date = request.form.get('pm_start_date', '')
        pm_start_time = request.form.get('pm_start_time', '')
        pm_end_date = request.form.get('pm_end_date', '')
        pm_end_time = request.form.get('pm_end_time', '')
        
        # Handle customer name
        customer = request.form.get('customer', '')
        if not customer:
            customer = request.form.get('customer_other', '')
        
        machine_type = request.form.get('machine_type', '')
        machine_model = request.form.get('machine_model', '')
        machine_serial = request.form.get('machine_serial', '')
        controller = request.form.get('controller', '')
        abnormalities = request.form.get('abnormalities', '')
        first_engineer = request.form.get('first_engineer', '')
        report_no = request.form.get('report_no', '')
        backup_status = request.form.get('backup_status', '')

        # Validate required fields
        required_fields = [pm_start_date, pm_start_time, pm_end_date, pm_end_time, 
                          customer, machine_type, machine_model, machine_serial, abnormalities, 
                          first_engineer, report_no, backup_status]
        
        if not all(required_fields):
            return jsonify({'success': False, 'message': 'Please fill all required fields!'})

        # Handle file uploads
        uploaded_files = []
        
        # Save photos
        photos = request.files.getlist('photos')
        for photo in photos:
            if photo and allowed_file(photo.filename, ALLOWED_IMAGES):
                filename = f"{uuid.uuid4().hex}_{secure_filename(photo.filename)}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                photo.save(filepath)
                uploaded_files.append({'type': 'photo', 'filename': filename, 'path': filepath})
        
        # Save videos
        videos = request.files.getlist('videos')
        for video in videos:
            if video and allowed_file(video.filename, ALLOWED_VIDEOS):
                filename = f"{uuid.uuid4().hex}_{secure_filename(video.filename)}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                video.save(filepath)
                uploaded_files.append({'type': 'video', 'filename': filename, 'path': filepath})

        # Combine date and time
        pm_start = f"{pm_start_date} {pm_start_time}" if pm_start_date and pm_start_time else ''
        pm_end = f"{pm_end_date} {pm_end_time}" if pm_end_date and pm_end_time else ''

        # Excel file operations
        file = "pm_data.xlsx"
        
        if os.path.exists(file):
            wb = load_workbook(file)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
        
        # Add headers if empty
        if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
            headers = [
                'Report No', 'PM Start', 'PM End', 'Customer', 'Machine Type' 
                'Machine Model', 'Machine Serial', 'Controller', 
                'Abnormalities Corrected', 'First Engineer', 'Backup Status',
                'Uploaded Files', 'Submission Date'
            ]
            sheet.append(headers)
        
        # Prepare file list string
        file_list = ', '.join([f"{f['type']}: {f['filename']}" for f in uploaded_files]) if uploaded_files else ''
        
        # Append data
        row_data = [
            pm_start,
            pm_end,
            customer,
            machine_type,
            machine_model,
            machine_serial,
            controller,
            abnormalities,
            first_engineer,
            report_no,
            backup_status,
            file_list,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        sheet.append(row_data)
        wb.save(file)
        wb.close()
        
        return jsonify({'success': True, 'message': 'Data saved successfully with files!'})
    
    except Exception as e:
        print(f"Error saving data: {str(e)}")
        return jsonify({'success': False, 'message': f'Error saving data: {str(e)}'})

if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)
